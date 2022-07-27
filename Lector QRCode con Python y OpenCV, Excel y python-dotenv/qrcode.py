import cv2
import numpy as np
import sys
from openpyxl import load_workbook
from settings import FILE_PATH, SHEET, COLUMN, ROW


wb = load_workbook(FILE_PATH)
ws = wb[SHEET]

capture = cv2.VideoCapture(0)

while(capture.isOpened()):
    ret, frame = capture.read()

    if (cv2.waitKey(1) == ord('s')):
        break
    qrDetector = cv2.QRCodeDetector()
    data, bbox, rectifiedImage = qrDetector.detectAndDecode(frame)

    if len(data) > 0:
        save = ''
        while save not in ['y', 'n']:
            save = input(f'El dato leido es {data}, quiere guardarlo [y/n]: ')
            if save == 'y':
                ws[f'{COLUMN}{ROW}'] = data
                ROW += 1
        cv2.imshow('webCam', rectifiedImage)
    else:
        cv2.imshow('webCam', frame)

wb.save(FILE_PATH)
capture.release()
cv2.destroyAllWindows()












