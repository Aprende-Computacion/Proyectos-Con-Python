import sys
from openpyxl import load_workbook

_, file_path, sheet, row = sys.argv
row = int(row)

wb = load_workbook(file_path)

ws = wb[sheet]

ws[f'B{row}'] = 123

wb.save(file_path)
